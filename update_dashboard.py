#!/usr/bin/env python3
"""
Andlz Dashboard — Gerador automático v2
Corrige bugs identificados: ROA decimal, CRR sublinhas, DCF formatação.

Uso:
    python update_dashboard_v2.py [planilha.xlsx]

Requisitos:
    pip install openpyxl
"""

import json, math, os, re, sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl --quiet")
    import openpyxl

SCRIPT_DIR = Path(__file__).parent
TEMPLATE   = SCRIPT_DIR / "dashboard_template.html"
OUTPUT     = SCRIPT_DIR / "dashboard_andlz.html"

if len(sys.argv) > 1:
    PLANILHA = Path(sys.argv[1])
else:
    xlsx_files = sorted(SCRIPT_DIR.glob("*.xlsx"))
    if not xlsx_files:
        print("ERRO: Nenhum .xlsx encontrado.")
        sys.exit(1)
    PLANILHA = xlsx_files[0]
    print(f"Planilha: {PLANILHA.name}")

MONTHS = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']

# ── HELPERS ───────────────────────────────────────────────────
def load_wb(path):
    return openpyxl.load_workbook(path, data_only=True)

def cell_float(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if v is None: return 0.0
    if isinstance(v, str):
        v = v.strip().replace('.','').replace(',','.')
        try: return float(v)
        except: return 0.0
    try: return float(v)
    except: return 0.0

def row_vals(ws, row, start_col, n=12):
    return [cell_float(ws, row, start_col + i) for i in range(n)]

def find_row(ws, label, label_col=1, max_rows=250, startswith=True):
    lc = str(label).strip().lower()
    for r in range(1, max_rows + 1):
        v = ws.cell(row=r, column=label_col).value
        if not v: continue
        vs = str(v).strip().lower()
        if startswith:
            if vs.startswith(lc[:20]): return r
        else:
            if vs == lc: return r
    return None

def arr12(vals, n):
    r = list(vals[:n])
    r += [0] * (12 - len(r))
    return r

def arr12f(vals):
    r = list(vals[:12])
    r += [0] * (12 - len(r))
    return r

def fmt_brl(v):
    """Format number as Brazilian currency string."""
    if v is None: return '–'
    n = abs(v)
    if n >= 1e6:   s = f"R$ {n/1e6:.1f}M"
    elif n >= 1e3: s = f"R$ {n/1e3:.0f}K"
    else:          s = f"R$ {n:,.0f}".replace(',','.')
    return ('-' if v < 0 else '') + s

def fmt_brl_full(v):
    """Full Brazilian currency string e.g. R$ 1.285.809"""
    if v is None: return '–'
    n = abs(v)
    s = f"R$ {n:,.0f}".replace(',','X').replace('.','.',).replace('X','.')
    return ('-' if v < 0 else '') + s

def fmt_pct(v, decimals=1):
    return f"{v*100:.{decimals}f}%"

# ── JS HELPERS ────────────────────────────────────────────────
def js_arr(lst):
    items = []
    for v in lst:
        if v == 'null' or v is None:
            items.append('null')
        elif isinstance(v, float) and math.isnan(v):
            items.append('0')
        else:
            items.append(str(round(float(v), 6)) if isinstance(v, float) else str(v))
    return '[' + ','.join(items) + ']'

def js_val(v):
    if v is None or v == 'null': return 'null'
    if isinstance(v, (int, float)): return str(round(v, 6))
    return json.dumps(v, ensure_ascii=False)

# ── LOAD WORKBOOK ─────────────────────────────────────────────
print(f"\nLendo {PLANILHA.name}...")
wb = load_wb(PLANILHA)
sheets = wb.sheetnames
print(f"Abas: {', '.join(sheets)}")

# ── CONSOLIDATED ──────────────────────────────────────────────
def read_consolidated():
    ws = wb['Consolidated']
    LC = 4   # label col D
    DC = 5   # data col E = Jan

    def get(label):
        r = find_row(ws, label, LC)
        return row_vals(ws, r, DC) if r else [0]*12

    rec = get('01 Receitas Ope')
    n   = sum(1 for v in rec if v != 0)
    print(f"Meses realizados: {n} ({', '.join(MONTHS[:n])})")

    # ROA — multiply by 100 to convert from decimal to percentage
    r_roa = find_row(ws, 'ROA Anual', LC)
    real_roa = [round(cell_float(ws, r_roa, DC+i) * 100, 4) for i in range(n)] if r_roa else [0]*n

    # AuM
    r_aum = find_row(ws, 'AuM', LC, startswith=False)
    if not r_aum:
        r_aum = find_row(ws, 'AuM', LC)
    real_aum = [int(cell_float(ws, r_aum, DC+i)) for i in range(n)] if r_aum else [0]*n

    # Caixa realizado
    r_caixa = find_row(ws, 'Caixa', LC, startswith=False)
    if not r_caixa:
        r_caixa = find_row(ws, 'Caixa', LC)
    real_caixa = [cell_float(ws, r_caixa, DC+i) for i in range(n)] if r_caixa else [0]*n

    # Caixa projetado
    r_cp = find_row(ws, 'Caixa projetado', LC)
    proj_caixa = arr12f(row_vals(ws, r_cp, DC) if r_cp else [0]*12)

    data = {
        'recBruta':    arr12(rec, n),
        'consultoria': arr12(get('Consultoria'), n),
        'planejamento':arr12(get('Planejamento'), n),
        'tesouroMT':   arr12(get('Tesouro Mato Grosso'), n),
        'rpps':        arr12(get('RPPS'), n),
        'consulta':    arr12(get('Consulta'), n),
        'valuation':   arr12(get('Valuation'), n),
        'cambio':      arr12(get('Câmbio/Bancários'), n),
        'consorcio':   arr12(get('Consórcio'), n),
        'previdencia': arr12(get('Previdência'), n),
        'segVida':     arr12(get('Seguro de vida'), n),
        'outrosRamos': arr12(get('Outros Ramos'), n),
        'deducoes':    arr12(get('02 Deduções'), n),
        'imp_vendas':  arr12(get('02.1 Impostos'), n),
        'com_consult': arr12(get('02.2 Comissões Cons'), n),
        'com_mesa':    arr12(get('02.3 Comissões Mesa'), n),
        'descontos':   arr12(get('02.4 Descontos'), n),
        'devolucoes':  arr12(get('02.5 Devoluções'), n),
        'com_parc':    arr12(get('02.6 Comissões Parc'), n),
        'recLiq':      arr12(get('02T Receita Líquida'), n),
        'custos':      arr12(get('03 Custos Oper'), n),
        'sal_enc':     arr12(get('03.1 Salários'), n),
        'sistemas_c':  arr12(get('03.2 Sistemas'), n),
        'custo_MT':    arr12(get('03.3 Custo Tesouro'), n),
        'custo_rpps':  arr12(get('03.4 Custo RPPS'), n),
        'regulatorias':arr12(get('03.5 Regulatórias'), n),
        'lucroBruto':  arr12(get('03T Lucro Bruto'), n),
        'despOp':      arr12(get('04 Despesas Oper'), n),
        'imovel':      arr12(get('04.1 Imóvel'), n),
        'servicos':    arr12(get('04.2 Serviços'), n),
        'comerciais':  arr12(get('04.3 Comerciais'), n),
        'admin':       arr12(get('04.4 Admin'), n),
        'colabor':     arr12(get('04.5 Colaboradores'), n),
        'diretoria':   arr12(get('04.6 Diretoria'), n),
        'lucroOp':     arr12(get('04T Lucro / Prejuízo Oper'), n),
        'financeiras': arr12(get('05 Receitas e Despesas Fin'), n),
        'outras_nd':   arr12(get('06 Outras'), n),
        'lucroLiq':    arr12(get('06T Lucro / Prejuízo Líq'), n),
        'inv_emp':     arr12(get('07 Despesas com Invest'), n),
        'inv_imob':    arr12(get('07.1 Investimentos em Imob'), n),
        'emprestimos': arr12(get('07.2 Empréstimos'), n),
        'lucroFinal':  arr12(get('07T Lucro / Prejuízo Final'), n),
        'realCaixa':   real_caixa,
        'realROA':     real_roa,
        'realAuM':     real_aum,
        'projCaixa':   proj_caixa,
    }

    # EBITDA = lucroOp + |imp_vendas|
    data['ebitda'] = [data['lucroOp'][i] + abs(data['imp_vendas'][i]) for i in range(12)]

    # Margens
    data['margEBITDA'] = [round(data['ebitda'][i]/data['recBruta'][i]*100,4) if data['recBruta'][i] else 0 for i in range(n)]
    data['margBruta']  = [round(data['lucroBruto'][i]/data['recBruta'][i]*100,4) if data['recBruta'][i] else 0 for i in range(n)]
    data['margLiq']    = [round(data['lucroLiq'][i]/data['recBruta'][i]*100,4) if data['recBruta'][i] else 0 for i in range(n)]
    data['margFinal']  = [round(data['lucroFinal'][i]/data['recBruta'][i]*100,4) if data['recBruta'][i] else 0 for i in range(n)]

    return data, n

# ── ORÇAMENTO 2026 ────────────────────────────────────────────
def read_orcamento(n):
    ws = wb['Orçamento 2026']
    LC, DC = 2, 3

    def get(label):
        r = find_row(ws, label, LC)
        return row_vals(ws, r, DC) if r else [0]*12

    orc = {
        'orcRecBruta':   arr12f(get('01 Receitas Ope')),
        'orcDeducoes':   arr12f(get('02 Deduções')),
        'orcCustos':     arr12f(get('03 Custos Oper')),
        'orcDespOp':     arr12f(get('04 Despesas Oper')),
        'orcLucroOp':    arr12f(get('04T Lucro / Prejuízo Oper')),
        'orcFinal':      arr12f(get('07T Lucro / Prejuízo Final')),
    }

    r_marg = find_row(ws, 'Margem EBITDA', LC)
    if r_marg:
        raw = row_vals(ws, r_marg, DC)
        orc['orcMargEBITDA'] = [round(v*100,4) if abs(v)<2 else round(v,4) for v in raw]
    else:
        orc['orcMargEBITDA'] = [0]*12

    # Meta AuM (célula C5)
    meta = ws.cell(row=5, column=3).value
    orc['metaAuM'] = int(float(meta)) if meta else 280000000

    orc['projFinal'] = ['null']*n + orc['orcFinal'][n:]

    # projAuM
    r_pa = find_row(ws, 'AuM Projetado', LC)
    proj_aum = ['null']*n
    if r_pa:
        for i in range(n, 12):
            v = cell_float(ws, r_pa, DC+i)
            proj_aum.append(int(v) if v else 'null')
    else:
        proj_aum += ['null']*(12-n)
    orc['projAuM'] = proj_aum

    return orc

# ── PROJEÇÃO (linha 62 — Caixa contínuo) ─────────────────────
def read_projecao():
    ws = wb['Projeção']
    # Search cols D (4) and A (1) for "Caixa" row with 12 values
    for lc, dc in [(4,5),(1,2)]:
        for r in range(58, 68):
            v = ws.cell(row=r, column=lc).value
            if not v or 'caixa' not in str(v).lower(): continue
            vals = row_vals(ws, r, dc, 12)
            if sum(1 for x in vals if x != 0) >= 5:
                print(f"  Caixa Projeção: linha {r} (label col {lc})")
                return arr12f(vals)
    return [0]*12

# ── AuM HISTÓRICO ─────────────────────────────────────────────
def read_aum_hist(n):
    ws = wb['AuM']
    series = {}
    mes_map = {'janeiro':0,'fevereiro':1,'março':2,'abril':3,'maio':4,'junho':5,
               'julho':6,'agosto':7,'setembro':8,'outubro':9,'novembro':10,'dezembro':11}
    for r in range(1, 500):
        ano = ws.cell(r,1).value
        mes = ws.cell(r,2).value
        val = ws.cell(r,3).value
        if not ano or not mes or not val: continue
        try:
            ano_i = int(ano); val_f = float(val)
        except: continue
        if ano_i < 2018 or ano_i > 2030: continue
        mi = mes_map.get(str(mes).lower().strip())
        if mi is not None:
            series.setdefault(ano_i, {})[mi] = val_f

    # Detect unit (R$ or millions)
    sample = [v for d in series.values() for v in d.values()]
    in_reais = any(v > 1e6 for v in sample[:10]) if sample else False

    aum_raw, hist_lbl, hist_val = [], [], []
    for yr in sorted(series):
        d = series[yr]
        nm = max(d)+1 if d else 0
        m = [(d[i]/1e6 if in_reais else d[i]) if i in d else 0 for i in range(nm)]
        if m: aum_raw.append({'y':yr,'m':m})
        if yr < 2026 and 11 in d:
            hist_lbl.append(str(yr))
            hist_val.append(d[11] if not in_reais else int(d[11]))

    real_aum_2026 = []
    if 2026 in series:
        d = series[2026]
        for i in range(n):
            if i in d:
                real_aum_2026.append(int(d[i]) if not in_reais else int(d[i]))

    return aum_raw, hist_lbl, hist_val, real_aum_2026

# ── CNS ───────────────────────────────────────────────────────
def read_cns(n):
    ws = wb['CNS']
    def get(label): r = find_row(ws,label,1); return arr12(row_vals(ws,r,2) if r else [0]*12, n)
    return {
        'cnsRB':      get('01 Receitas Operacionais'),
        'cnsRL':      get('02T Receita Líquida'),
        'cnsFinal':   get('07T Lucro / Prejuízo Final'),
        'cnsConsult': get('Consultoria'),
        'cnsPlan':    get('Planejamento'),
        'cnsMT':      get('Tesouro Mato Grosso'),
        'cnsRPPS':    get('RPPS'),
        'cnsConsulta':get('Consulta'),
    }

# ── CRR ───────────────────────────────────────────────────────
def read_crr(n):
    ws = wb['CRR']
    # Labels in col 1, data starts col 2
    # Sublines use prefixes like "01.1 Câmbio/Bancários"
    def get(label): r = find_row(ws,label,1); return arr12(row_vals(ws,r,2) if r else [0]*12, n)
    return {
        'crrRB':       get('01 Receitas Operacionais'),
        'crrRL':       get('02T Receita Líquida'),
        'crrFinal':    get('07T Lucro / Prejuízo Final'),
        'crrCambio':   get('01.1 Câmbio'),          # FIX: prefix 01.1
        'crrConsorcio':get('01.2 Consórcio'),        # FIX: prefix 01.2
        'crrPrevid':   get('01.3 Previdência'),      # FIX: prefix 01.3 (or try without)
        'crrSegVida':  get('01.4 Seguro'),           # FIX: prefix 01.4
    }

# ── RECEITA (consultores) ─────────────────────────────────────
def read_receita(n):
    ws = wb['Receita']
    cons = {}
    for r in range(15, 600):
        nome = ws.cell(r,9).value
        if not nome or not isinstance(nome,str) or len(nome.strip())<2: continue
        nome = nome.strip()
        vals = [cell_float(ws,r,12+i) for i in range(12)]
        if not any(vals[:n]): continue
        if nome not in cons: cons[nome] = [0.0]*12
        for i in range(12): cons[nome][i] += vals[i]
    result = [{'n':k,'v':arr12(v,n)} for k,v in cons.items()]
    result.sort(key=lambda x: sum(x['v']), reverse=True)
    return result

# ── GASTOS ────────────────────────────────────────────────────
def read_gastos(n):
    ws = wb['Consolidated']
    LC, DC = 4, 5
    def get_abs(label):
        r = find_row(ws, label, LC)
        return [abs(cell_float(ws,r,DC+i)) for i in range(12)] if r else [0]*12
    cost_map = [
        ('Salários e Encargos', '03.1 Salários',      False),
        ('Mato Grosso',         '03.3 Custo Tesouro',  False),
        ('Sistemas',            '03.2 Sistemas',       False),
        ('Desp. Operacionais',  '04 Despesas Oper',    False),
        ('Financeiras',         '05 Receitas e Desp',  False),
        ('Regulatórias',        '03.5 Regulatórias',   False),
        ('Comissões',           '02.2 Comissões Cons', True),
        ('Impostos s/ Vendas',  '02.1 Impostos',       True),
        ('PLR/Bônus',           None,                  True),
    ]
    gastos = []
    for nome, label, excl in cost_map:
        vals = get_abs(label) if label else [0]*12
        if sum(vals[:n]) > 0:
            gastos.append({'n':nome,'v':arr12(vals,n),'excl':excl})
    return gastos

# ── SISTEMAS ──────────────────────────────────────────────────
def read_sistemas():
    wn = 'Sistemas ' if 'Sistemas ' in sheets else 'Sistemas'
    if wn not in sheets: return [], 255
    ws = wb[wn]
    sistemas, cat = [], 'Geral'
    for r in range(1,100):
        cat_c = ws.cell(r,5).value
        nome  = ws.cell(r,6).value
        val   = ws.cell(r,7).value
        if cat_c and isinstance(cat_c,str) and not nome:
            cat = cat_c.strip(); continue
        if nome and isinstance(nome,str) and val is not None:
            nome = nome.strip()
            if nome.lower() in ['total','subtotal'] or nome.isdigit(): continue
            try:
                v = float(val)
                if v > 0: sistemas.append({'n':nome,'cat':cat,'v':v})
            except: pass
    return sistemas, 255

# ── H. RECEITA ────────────────────────────────────────────────
def read_h_receita():
    ws = wb['H. Receita']
    from datetime import datetime
    qmap={1:'1T',2:'1T',3:'1T',4:'2T',5:'2T',6:'2T',7:'3T',8:'3T',9:'3T',10:'4T',11:'4T',12:'4T'}
    qacc = {}
    for r in range(4,600):
        dt_v = ws.cell(r,4).value
        lucro= ws.cell(r,6).value
        if not dt_v or lucro is None: continue
        try:
            dt = dt_v if isinstance(dt_v, datetime) else datetime.strptime(str(dt_v)[:10],'%Y-%m-%d')
            key = qmap[dt.month]+str(dt.year)[-2:]
            qacc[key] = qacc.get(key,0)+float(lucro or 0)
        except: continue
    lbl = sorted(qacc.keys())
    return lbl, [round(qacc[k],2) for k in lbl]

# ── CAPTAÇÃO ──────────────────────────────────────────────────
def read_captacao(n):
    if 'Captação' not in sheets: return None
    ws = wb['Captação']
    LC, DC = 6, 7  # label col F, data col G
    def row_arr(row_num): return arr12(row_vals(ws,row_num,DC), n)
    return {
        'numClientes':   row_arr(5),
        'captLiquida':   row_arr(6),
        'ticketMedio':   row_arr(7),
        'saldoClientes': row_arr(8),
        'captBruta':     row_arr(9),
        'novosContratos':row_arr(10),
        'renovacoes':    row_arr(11),
        'entradaClientes':row_arr(12),
        'churnTotal':    row_arr(13),
        'retiradas':     row_arr(14),
        'cancelamento':  row_arr(15),
        'saidaClientes': row_arr(16),
    }

# ── DCF ───────────────────────────────────────────────────────
def read_dcf():
    if 'DCF Andlz' not in sheets: return None
    ws = wb['DCF Andlz']
    anos = ['2027','2028','2029','2030','2031','Perpetuidade']

    # Premissas B4-B8
    ebitda_base = cell_float(ws,4,2)
    wacc        = cell_float(ws,5,2)
    tx_cresc    = cell_float(ws,6,2)
    tx_perp     = cell_float(ws,7,2)
    pct_fcf     = cell_float(ws,8,2)

    # Tabela FCF — find header row (row 11 has "Item", "2027"...)
    # Data rows 12-15
    fcf_rows = []
    for r in range(12, 18):
        item = ws.cell(r,1).value
        if not item: continue
        item = str(item).strip()
        if item.upper() in ['PROJEÇÃO DE FLUXO DE CAIXA LIVRE','ITEM']: continue
        row_data = {'Item': item}
        for i, ano in enumerate(anos):
            v = ws.cell(r, 2+i).value
            if v is None or v == '–':
                row_data[ano] = '–'
            else:
                try:
                    fv = float(v)
                    # Format according to row type
                    if 'fator' in item.lower() or 'taxa' in item.lower():
                        row_data[ano] = f"{fv:.4f}"
                    elif 'desconto' in item.lower() and fv < 2:
                        row_data[ano] = f"{fv:.4f}"
                    else:
                        # Currency — format as "R$ 1.285.809"
                        row_data[ano] = fmt_brl_full(fv)
                except:
                    row_data[ano] = str(v)
        fcf_rows.append(row_data)

    # Tabela Valuation rows 19-24
    val_rows = []
    for r in range(19, 26):
        item = ws.cell(r,1).value
        val  = ws.cell(r,2).value
        if not item: continue
        item = str(item).strip()
        if item.upper() in ['VALUATION','ITEM']: continue
        try:
            fv = float(val)
            val_str = fmt_brl_full(fv)
        except:
            val_str = str(val) if val else '–'
        val_rows.append({'Item': item, 'Valor': val_str})

    return {
        'ebitdaBase':    ebitda_base,
        'wacc':          wacc,
        'txCrescAnual':  tx_cresc,
        'txCrescPerp':   tx_perp,
        'pctFCF':        pct_fcf,
        'fcfRows':       fcf_rows,
        'valRows':       val_rows,
    }

# ── EXECUÇÃO ──────────────────────────────────────────────────
print("\nExtraindo dados...")
consolidated, n = read_consolidated()
orcamento       = read_orcamento(n)
proj_caixa62    = read_projecao()
aum_raw, hist_lbl, hist_val, real_aum_2026 = read_aum_hist(n)
cns             = read_cns(n)
crr             = read_crr(n)
consultores     = read_receita(n)
gastos          = read_gastos(n)
sistemas, num_cli = read_sistemas()
hist_labels, hist_vals = read_h_receita()
captacao        = read_captacao(n)
dcf             = read_dcf()

# Merge realAuM (prefer Consolidated, fallback to AuM sheet)
real_aum = consolidated.get('realAuM') or real_aum_2026

print(f"  Consultores: {len(consultores)}")
print(f"  Sistemas: {len(sistemas)}")
print(f"  Captação: {'✓' if captacao else '–'}")
print(f"  DCF: {'✓' if dcf else '–'}")
print(f"  ROA[0]: {consolidated['realROA'][0] if consolidated['realROA'] else '–'}")

# ── MONTA RAW ─────────────────────────────────────────────────
def aum_series_js(s):
    lines = ['[\n']
    for yr in s:
        m = ','.join(str(round(v,2)) for v in yr['m'])
        lines.append(f"    {{y:{yr['y']},m:[{m}]}},\n")
    lines.append('  ]')
    return ''.join(lines)

def cons_js(lst):
    lines = ['[\n']
    for c in lst:
        lines.append(f"    {{n:{json.dumps(c['n'])},v:{js_arr(c['v'])}}},\n")
    lines.append('  ]')
    return ''.join(lines)

def gastos_js(lst):
    lines = ['[\n']
    for g in lst:
        e = 'true' if g['excl'] else 'false'
        lines.append(f"    {{n:{json.dumps(g['n'])},v:{js_arr(g['v'])},excl:{e}}},\n")
    lines.append('  ]')
    return ''.join(lines)

def sistemas_js(lst):
    lines = ['[\n']
    for s in lst:
        lines.append(f"    {{n:{json.dumps(s['n'])},cat:{json.dumps(s['cat'])},v:{s['v']}}},\n")
    lines.append('  ]')
    return ''.join(lines)

def captacao_js(c):
    if not c:
        z = js_arr([0]*12)
        fields = ['numClientes','captLiquida','ticketMedio','saldoClientes',
                  'captBruta','novosContratos','renovacoes','entradaClientes',
                  'churnTotal','retiradas','cancelamento','saidaClientes']
        return '{\n' + ''.join(f"    {f}:{z},\n" for f in fields) + '  }'
    fields = ['numClientes','captLiquida','ticketMedio','saldoClientes',
              'captBruta','novosContratos','renovacoes','entradaClientes',
              'churnTotal','retiradas','cancelamento','saidaClientes']
    return '{\n' + ''.join(f"    {f}:{js_arr(c[f])},\n" for f in fields) + '  }'

def dcf_js(d):
    if not d:
        return '{ebitdaBase:null,wacc:null,txCrescAnual:null,txCrescPerp:null,pctFCF:null,fcfRows:[],valRows:[]}'
    fcf = json.dumps(d['fcfRows'], ensure_ascii=False)
    val = json.dumps(d['valRows'], ensure_ascii=False)
    return (f'{{\n    ebitdaBase:{js_val(d["ebitdaBase"])},\n'
            f'    wacc:{js_val(d["wacc"])},\n'
            f'    txCrescAnual:{js_val(d["txCrescAnual"])},\n'
            f'    txCrescPerp:{js_val(d["txCrescPerp"])},\n'
            f'    pctFCF:{js_val(d["pctFCF"])},\n'
            f'    fcfRows:{fcf},\n'
            f'    valRows:{val},\n  }}')

proj_final_js = '[' + ','.join('null' if v=='null' else str(round(float(v),6)) for v in orcamento['projFinal']) + ']'
proj_aum_js   = '[' + ','.join('null' if v=='null' else str(v) for v in orcamento['projAuM']) + ']'

raw_block = f"""const RAW = {{
  recBruta:{js_arr(consolidated['recBruta'])},
  consultoria:{js_arr(consolidated['consultoria'])},
  planejamento:{js_arr(consolidated['planejamento'])},
  tesouroMT:{js_arr(consolidated['tesouroMT'])},
  rpps:{js_arr(consolidated['rpps'])},
  consulta:{js_arr(consolidated['consulta'])},
  valuation:{js_arr(consolidated['valuation'])},
  cambio:{js_arr(consolidated['cambio'])},
  consorcio:{js_arr(consolidated['consorcio'])},
  previdencia:{js_arr(consolidated['previdencia'])},
  segVida:{js_arr(consolidated['segVida'])},
  outrosRamos:{js_arr(consolidated['outrosRamos'])},
  deducoes:{js_arr(consolidated['deducoes'])},
  imp_vendas:{js_arr(consolidated['imp_vendas'])},
  com_consult:{js_arr(consolidated['com_consult'])},
  com_mesa:{js_arr(consolidated['com_mesa'])},
  descontos:{js_arr(consolidated['descontos'])},
  devolucoes:{js_arr(consolidated['devolucoes'])},
  com_parc:{js_arr(consolidated['com_parc'])},
  recLiq:{js_arr(consolidated['recLiq'])},
  custos:{js_arr(consolidated['custos'])},
  sal_enc:{js_arr(consolidated['sal_enc'])},
  sistemas_c:{js_arr(consolidated['sistemas_c'])},
  custo_MT:{js_arr(consolidated['custo_MT'])},
  custo_rpps:{js_arr(consolidated['custo_rpss'] if 'custo_rpss' in consolidated else consolidated['custo_rpps'])},
  regulatorias:{js_arr(consolidated['regulatorias'])},
  lucroBruto:{js_arr(consolidated['lucroBruto'])},
  despOp:{js_arr(consolidated['despOp'])},
  imovel:{js_arr(consolidated['imovel'])},
  servicos:{js_arr(consolidated['servicos'])},
  comerciais:{js_arr(consolidated['comerciais'])},
  admin:{js_arr(consolidated['admin'])},
  colabor:{js_arr(consolidated['colabor'])},
  diretoria:{js_arr(consolidated['diretoria'])},
  lucroOp:{js_arr(consolidated['lucroOp'])},
  financeiras:{js_arr(consolidated['financeiras'])},
  outras_nd:{js_arr(consolidated['outras_nd'])},
  lucroLiq:{js_arr(consolidated['lucroLiq'])},
  inv_emp:{js_arr(consolidated['inv_emp'])},
  inv_imob:{js_arr(consolidated['inv_imob'])},
  emprestimos:{js_arr(consolidated['emprestimos'])},
  lucroFinal:{js_arr(consolidated['lucroFinal'])},
  ebitda:{js_arr(consolidated['ebitda'])},
  margEBITDA:{js_arr(consolidated['margEBITDA'])},
  margBruta:{js_arr(consolidated['margBruta'])},
  margLiq:{js_arr(consolidated['margLiq'])},
  margFinal:{js_arr(consolidated['margFinal'])},
  projFinal:{proj_final_js},
  projCaixa:{js_arr(consolidated['projCaixa'])},
  projCaixaLinha62:{js_arr(proj_caixa62)},
  realAuM:{js_arr(real_aum)},
  projAuM:{proj_aum_js},
  metaAuM:{orcamento['metaAuM']},
  realROA:{js_arr(consolidated['realROA'])},
  realCaixa:{js_arr(consolidated['realCaixa'])},
  orcRecBruta:{js_arr(orcamento['orcRecBruta'])},
  orcDeducoes:{js_arr(orcamento['orcDeducoes'])},
  orcCustos:{js_arr(orcamento['orcCustos'])},
  orcDespOp:{js_arr(orcamento['orcDespOp'])},
  orcLucroOp:{js_arr(orcamento['orcLucroOp'])},
  orcFinal:{js_arr(orcamento['orcFinal'])},
  orcMargEBITDA:{js_arr(orcamento['orcMargEBITDA'])},
  cnsRB:{js_arr(cns['cnsRB'])},
  cnsRL:{js_arr(cns['cnsRL'])},
  cnsFinal:{js_arr(cns['cnsFinal'])},
  cnsConsult:{js_arr(cns['cnsConsult'])},
  cnsPlan:{js_arr(cns['cnsPlan'])},
  cnsMT:{js_arr(cns['cnsMT'])},
  cnsRPPS:{js_arr(cns['cnsRPPS'])},
  cnsConsulta:{js_arr(cns['cnsConsulta'])},
  crrRB:{js_arr(crr['crrRB'])},
  crrRL:{js_arr(crr['crrRL'])},
  crrFinal:{js_arr(crr['crrFinal'])},
  crrCambio:{js_arr(crr['crrCambio'])},
  crrConsorcio:{js_arr(crr['crrConsorcio'])},
  crrPrevid:{js_arr(crr['crrPrevid'])},
  crrSegVida:{js_arr(crr['crrSegVida'])},
  consultores:{cons_js(consultores)},
  gastosAll:{gastos_js(gastos)},
  histLabels:{json.dumps(hist_labels)},
  histVals:{js_arr(hist_vals)},
  histRecLabels:[],
  histRecVals:[],
  aumHistLabels:{json.dumps(hist_lbl)},
  aumHistVals:{js_arr(hist_val)},
  aumRawSeries:{aum_series_js(aum_raw)},
  sistemas:{sistemas_js(sistemas)},
  numClientes:{num_cli},
  captacao:{captacao_js(captacao)},
  dcf:{dcf_js(dcf)},
}};"""

# ── INJETA NO TEMPLATE ────────────────────────────────────────
print("\nLendo template...")
tmpl = TEMPLATE.read_text(encoding='utf-8')

marker = "// <<SKILL_DATA_INJECTION>>\nconst RAW = __INJECTED_DATA__;"
if marker not in tmpl:
    print("ERRO: marcador não encontrado no template.")
    sys.exit(1)

out = tmpl.replace(marker, raw_block)
out = re.sub(
    r'let pFrom = 0, pTo = __DEFAULT_PTO__;.*\n',
    f'let pFrom = 0, pTo = {n-1}; // {MONTHS[n-1]}/26\n',
    out
)

# Update period selectors
opts_to   = '\n'.join(f'        <option value="{i}"{" selected" if i==n-1 else ""}>{MONTHS[i]}</option>' for i in range(n))
opts_from = '\n'.join(f'        <option value="{i}">{MONTHS[i]}</option>' for i in range(n))
out = re.sub(r'<select class="period-select" id="sel-to">.*?</select>', f'<select class="period-select" id="sel-to">\n{opts_to}\n        </select>', out, flags=re.DOTALL)
out = re.sub(r'<select class="period-select" id="sel-from">.*?</select>', f'<select class="period-select" id="sel-from">\n{opts_from}\n        </select>', out, flags=re.DOTALL)

OUTPUT.write_text(out, encoding='utf-8')
kb = OUTPUT.stat().st_size/1024
print(f"\n✓ {OUTPUT.name} ({kb:.0f} KB)")
print(f"  Período: Jan–{MONTHS[n-1]}/2026 ({n} meses)")
print(f"  Meta AuM: R$ {orcamento['metaAuM']:,}")
print(f"  ROA[0]: {consolidated['realROA'][0] if consolidated['realROA'] else '?'}")
print(f"  CRR Câmbio Jan: {crr['crrCambio'][0] if crr['crrCambio'] else '?'}")
print(f"  DCF EBITDA: {dcf['ebitdaBase'] if dcf else '–'}")
